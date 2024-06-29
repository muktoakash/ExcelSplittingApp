"""./main.py

This is an app that splits up an Excel file into many based on the current sheets in the spreadsheet."""

# Imports
import os, sys, re
import win32com.client as win32
from openpyxl import load_workbook
from pywintypes import com_error
from PyQt6.QtCore import Qt
from PyQt6.QtGui import QIcon
from PyQt6.QtWidgets import (QApplication, QWidget, QPushButton, QListWidget, QLineEdit, QStatusBar, QListWidgetItem, QFileDialog, QFormLayout, QHBoxLayout, QVBoxLayout)

def get_worksheet_names(file_path : str):
    """
    Loads all the sheet names from a given file path.
    Return type: list?
    """
    workbook = load_workbook(file_path)
    return workbook.sheetnames

class AppWindow(QWidget):
    """"""
    def __init__(self):
        super().__init__()
        self.setWindowTitle('Excel File Splitter')
        self.window_width, self.window_height = 700, 50
        self.setMinimumSize(self.window_width, self.window_height)
        self.setWindowIcon(QIcon('icon.png'))
        self.setStyleSheet('''
                           QWidget {
                           font-size: 14px;
                           }
                           '''
                           )
        self.setObjectName('AppWindow')
        self.setLayout(self.layout['main'])

        self.initUI()
        self.config_signals()

        def initUI(self):
            self.layout['form'] = QFormLayout()
            self.layout['main'].addLayout(self.layout['form'])

            self.layout['browse_file'] = QHBoxLayout()

            self.file_path = QLineEdit()
            self.layout['browse_file'].addWidget(self.file_path)

            self.button_browse = QPushButton('Browse')
            self.layout['browse_file'].addWidget(self.button_browse)

            self.layout['form'].addRow('File Path: ', self.layout['browse_file'])

            self.instant_search = QLineEdit()
            self.layout['form'].addRow('Search: ' self.instant_search)

            self.list_sheet_name = QListWidget()
            self.layout['form'].addRow(self.list_sheet_name)

            self.button_split = QPushButton('Split')
            self.layout['form'].addRow(self.button_split)

            self.status_bar = QStatusBar()
            self.layout['main'].addWidget(self.status_bar)

    def browse_file(self):
        file_path, _ = QFileDialog.getOpenFileName(self, 'Open File', '', 'Excel Files (*.xlsx)')
        self.file_path.setText(file_path)

        if file_path:
            sheet_names = get_worksheet_names(file_path)

            self.list_sheet_name.clear()

            add_all_items = QListWidgetItem("Add All")
            add_all_items.setFlags(add_all_items.flags() | Qt.ItemFlag.ItemIsUserCheckable)
            add_all_items.setCheckState(Qt.CheckState.Unchecked)
            self.list_sheet_name.addItem(add_all_items)

            for sheet_name in sheet_names:
                listWidgetItem = QListWidget(sheet_name)
                listWidgetItem.setFlags(listWidgetItem.flags() | Qt.ItemFlag.ItemIsUserCheckable)
                listWidgetItem.setCheckState(Qt.CheckState.Unchecked)
                self.list_sheet_name.addItem(listWidgetItem)

            if len(sheet_names) == 1:
                self.status_bar.showMessage('Excel File has only one sheet. No need to split.')

    def check_uncheck_all(self, item):
        if item.text() == "Add All":
            for index in range(self.list_sheet_name.count()):
               list_box_item = self.list_Sheet_name.item(index)
               if list_box_item.text() != "Add All" and not list_box_item.isHidden():
                   list_box_item.setCheckState(item.checkState())

    def split_excel_file(self):
        file_path = self.file_path.text()

        if not file_path:
            self.status_bar.showMessage
            ('Please select an Excel file first.')

        if self.list_sheet_name.count() == 2:
                self.status_bar.showMessage('Excel file has only one sheet. No need to split.')
                return

        work_dir = os.getcwd()

        checked_sheets = []
        for index in range(self.list_sheet_name.count()):
            if self.list_sheet_name.item(index).text() != 'Add All' and self.list_sheet_name.item(index).checkState() == Qt.CheckState.Checked:
                checked_sheets.append(self.list_sheet_name.item(index).text())

        if not checked_sheets:
            self.status_bar.showMessage('Please select at least one sheet to split')
            return

        try:
            excel_app = win32.Dispatch('Excel.Application')
            excel_app.Visible = True
            workbook = excel_app.Workbooks.Open(file_path)

            for sheet_name in checked_sheets:
                new_workbook = excel_app.Workbooks.Add()
                workbook.Sheets(sheet_name).Copy(Before=new_workbook.Sheets(1))

                excel_app.DixplayAlerts = False
                new_workbook.Sheets(2).Delete()
                excel_app.DisplayAlerts = True

                valid_filename = re.sub(r'[\\/*?:<>|]', '_', sheet_name)
                new_workbook.SaveAs(os.path.join(work_dir, valid_filename + '.xlsx'))
                new_workbook.Close()

            workbook.Close(SaveChange = False)
            excel_app.Quit()

            self.status_bar.showMessage('Excel file has been split successfully.')

        except Exception as e:
            self.status_bar.showMessage(str(e))
            excel_app.DisplayAlerts = True
            workbook.Close(SaveChanges = False)
            excel_app.Quit()

    def fcn_instant_search(self):
        search_text = self.instant_search.text()
        if not search_text:
            for index in range(self.list_sheet_name.count()):
                self.list_sheet_name.item(index).setHidden(False)

        for index in range(self.list_sheet_name.count()):
            if search_text.lower() in self.list_sheet_name.item(index).text().lower or self.list_sheet_name.item(index).text() == "Add All":
                self.list_sheet_name.item(index).setHidden(True)
            else:
                self.list_sheet_name.item(index).setHidden(True)

# Run
if __name__ == '__main__':
    app = QApplication(sys.argv)
    app.setStyle('Fusion')
    app.setStyleSheet(open('style.css').read())

    myApp = AppWindow()
    myApp.show()

    try:
        sys.exit(app.exec())
    except SystemExit:
        print("Closing Window ... ")
